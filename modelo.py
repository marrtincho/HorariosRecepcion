"""
Hotel Staff Scheduler
Generates weekly schedules with shifts: Mañana, Tarde, Noche, Partido
Rules:
- Each person gets 2 consecutive days off per week
- night_fixed always off Thu+Fri; normal employees rotate covering those 2 nights
- Min/max coverage per shift per day
- Noche/Tarde cannot be followed by Mañana the next day (per person)
- Max 3 per Mañana/Tarde; max 1 per Noche/Partido per day
- jefe: Mañana obligado, no Noche, always off Sat+Sun
- subjefe: prefers Tarde, no Noche
- adjunto_recepcion: no Noche, consecutive days off normal
"""

import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
import copy

# ─── CONFIGURATION ──────────────────────────────────────────────────────────

SHIFTS = ["Mañana", "Tarde", "Noche", "Partido", "Libre"]
DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# Night-fixed always off these two days (Thu=3, Fri=4)
NIGHT_FIXED_DAYS_OFF = [3, 4]

MIN_COVERAGE = {"Mañana": 2, "Tarde": 2, "Noche": 1, "Partido": 0}
MAX_COVERAGE = {"Mañana": 3, "Tarde": 3, "Noche": 1, "Partido": 1}

NO_MORNING_AFTER = {"Noche", "Tarde"}

# Roles that can NEVER work Noche
NO_NIGHT_ROLES = {"jefe", "subjefe", "adjunto_recepcion"}

STAFF = [
    {"name": "Carlos", "role": "night_fixed", "preferences": ["Noche"]},
    {"name": "Juan", "role": "jefe", "preferences": ["Mañana"]},
    {"name": "Laura", "role": "subjefe", "preferences": ["Tarde"]},
    {"name": "Pedro", "role": "normal", "preferences": ["Mañana", "Partido"]},
    {"name": "Ana", "role": "normal", "preferences": []},
    {"name": "Miguel", "role": "normal", "preferences": ["Tarde"]},
    {"name": "Sofía", "role": "adjunto_recepcion", "preferences": ["Mañana"]},
    {"name": "Diego", "role": "normal", "preferences": []},
]


# ─── HELPER FUNCTIONS ────────────────────────────────────────────────────────

def get_consecutive_day_pairs():
    pairs = []
    for i in range(len(DAYS) - 1):
        pairs.append((i, i + 1))
    return pairs


def select_night_covers(staff_list: list, week_start: date) -> dict:
    """
    Selects ONE normal employee to cover Noche on BOTH Thu(3) and Fri(4),
    and assigns them Sat(5)+Sun(6) off as compensation.
    Skips employees who have forced_days_off or peticiones on Thu or Fri.
    Rotates fairly, preferring 'cobrar' employees.
    """
    # Employees with forced day off on Thu or Fri cannot cover those nights
    blocked = set()
    for p in staff_list:
        forced = p.get("forced_days_off", [])
        if "Jueves" in forced or "Viernes" in forced:
            blocked.add(p["name"])

    eligible = [
        p for p in staff_list
        if p["role"] == "normal" and p["name"] not in blocked
    ]
    if not eligible:
        return {}

    try:
        import json as _json, os as _os
        from herramientas import DATA_DIR
        hist_path = _os.path.join(DATA_DIR, "night_rotation.json")
        history = _json.loads(open(hist_path, encoding="utf-8").read()) \
            if _os.path.exists(hist_path) else {}
    except Exception:
        history = {}

    try:
        from herramientas import cargar_compensaciones
        comp = cargar_compensaciones()
    except Exception:
        comp = {}

    def sort_key(p):
        count = history.get(p["name"], 0)
        prefers_cobrar = 0 if comp.get(p["name"], {}).get("modalidad", "cobrar") == "cobrar" else 1
        return (count, prefers_cobrar, p["name"])

    sorted_eligible = sorted(eligible, key=sort_key)
    chosen_person = sorted_eligible[0]
    chosen_name = chosen_person["name"]

    chosen_person["forced_days_off"] = ["Sábado", "Domingo"]

    try:
        import json as _json, os as _os
        from herramientas import DATA_DIR
        history[chosen_name] = history.get(chosen_name, 0) + 1
        with open(_os.path.join(DATA_DIR, "night_rotation.json"), "w", encoding="utf-8") as f:
            _json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return {3: chosen_name, 4: chosen_name}


def _get_weekend_counts() -> dict:
    """
    Returns {name: count_of_weeks_with_both_Sat_and_Sun_libre}
    calculated from the historial, for balancing weekend days off.
    """
    try:
        import json as _json, os as _os
        from herramientas import DATA_DIR, cargar_historial
        historial = cargar_historial()
        counts = {}
        for entrada in historial.values():
            for nombre, turnos in entrada.get("schedule", {}).items():
                if len(turnos) >= 7 and turnos[5] == "Libre" and turnos[6] == "Libre":
                    counts[nombre] = counts.get(nombre, 0) + 1
        return counts
    except Exception:
        return {}


def assign_days_off(staff_list, night_cover: dict = None):
    """
    Assigns 2 consecutive days off + any extra LA days (non-replacing).
    - jefe:        always off Sáb(5)+Dom(6).
    - night_fixed: always off Thu(3)+Fri(4).
    - night cover person: forced Sáb+Dom (set by select_night_covers).
    - forced_days_off: from peticiones/vacaciones — used as regular days off.
    - extra_days_off:  LA days — added on top, don't replace the regular pair.
    - Normal staff: least-used consecutive pair, balancing weekends across history.
    """
    if night_cover is None:
        night_cover = {}

    cover_by_person = {}
    for day_idx, name in night_cover.items():
        cover_by_person.setdefault(name, []).append(day_idx)

    pairs = get_consecutive_day_pairs()
    schedule = {p["name"]: ["" for _ in DAYS] for p in staff_list}
    pair_usage = {pair: 0 for pair in pairs}

    # Weekend history for balancing (Sáb=5, Dom=6 pair)
    weekend_counts = _get_weekend_counts()

    jefe = next((p for p in staff_list if p["role"] == "jefe"), None)
    night_fixed = next((p for p in staff_list if p["role"] == "night_fixed"), None)

    # jefe: fixed Sáb+Dom off
    if jefe:
        schedule[jefe["name"]][5] = "Libre"
        schedule[jefe["name"]][6] = "Libre"
        jefe["days_off"] = [5, 6]

    # night_fixed: fixed Thu+Fri off
    if night_fixed:
        for d in NIGHT_FIXED_DAYS_OFF:
            schedule[night_fixed["name"]][d] = "Libre"
        night_fixed["days_off"] = list(NIGHT_FIXED_DAYS_OFF)
        pair_usage[tuple(NIGHT_FIXED_DAYS_OFF)] = \
            pair_usage.get(tuple(NIGHT_FIXED_DAYS_OFF), 0) + 1

    # Everyone else
    for person in staff_list:
        if person["role"] in ("jefe", "night_fixed"):
            continue

        forced = person.get("forced_days_off", [])
        forced_indices = [DAYS.index(d) for d in forced if d in DAYS]
        must_work = set(cover_by_person.get(person["name"], []))

        # ── Assign regular 2 consecutive days off ──────────────────────────
        if len(forced_indices) >= 2:
            chosen_idxs = forced_indices[:2]
            for idx in chosen_idxs:
                schedule[person["name"]][idx] = "Libre"
            person["days_off"] = sorted(chosen_idxs)
            pair = tuple(sorted(chosen_idxs))
            if pair in pair_usage:
                pair_usage[pair] += 1

        elif len(forced_indices) == 1:
            fi = forced_indices[0]
            candidates = [(fi - 1, fi)] if fi > 0 else []
            if fi < len(DAYS) - 1:
                candidates.append((fi, fi + 1))
            valid = [c for c in candidates if c in pair_usage and not (must_work & set(c))]
            if not valid:
                valid = [c for c in candidates if c in pair_usage]
            chosen = min(valid, key=lambda p: pair_usage[p]) if valid \
                else sorted(pairs, key=lambda p: pair_usage[p])[0]
            schedule[person["name"]][chosen[0]] = "Libre"
            schedule[person["name"]][chosen[1]] = "Libre"
            person["days_off"] = list(chosen)
            pair_usage[chosen] += 1

        else:
            # Balance weekends: prefer giving Sáb+Dom to those with fewest
            valid_pairs = [p for p in pairs if not (must_work & set(p))]
            if not valid_pairs:
                valid_pairs = pairs

            wc = weekend_counts.get(person["name"], 0)

            # Sort by: pair_usage first, but if weekend pair (5,6) hasn't been
            # used much by this person, give it slight priority
            def pair_sort_key(pair):
                usage = pair_usage[pair]
                # Bonus for weekend pair when this person hasn't had enough
                weekend_bonus = -0.5 if pair == (5, 6) and wc < 1 else 0
                return usage + weekend_bonus

            sorted_pairs = sorted(valid_pairs, key=pair_sort_key)
            chosen = sorted_pairs[0]
            pair_usage[chosen] += 1
            schedule[person["name"]][chosen[0]] = "Libre"
            schedule[person["name"]][chosen[1]] = "Libre"
            person["days_off"] = list(chosen)
            if chosen == (5, 6):
                weekend_counts[person["name"]] = wc + 1

        # ── Apply extra LA days on TOP (don't replace regular days off) ───
        extra = person.get("extra_days_off", [])
        for day_name in extra:
            if day_name in DAYS:
                idx = DAYS.index(day_name)
                schedule[person["name"]][idx] = "Libre"

    return schedule


def can_work_morning(person_name, day_idx, schedule):
    if day_idx == 0:
        return True
    prev = schedule[person_name][day_idx - 1]
    return prev not in NO_MORNING_AFTER


def allowed_shifts(person, day_idx, schedule, coverage, night_cover_today: bool = False):
    """
    Returns assignable day-shifts for a person (Mañana/Tarde/Partido).
    Night is handled separately in Pass 1.
    """
    result = []
    for shift in ["Mañana", "Tarde", "Partido"]:
        if coverage.get(shift, 0) >= MAX_COVERAGE.get(shift, 99):
            continue
        if shift == "Mañana" and not can_work_morning(person["name"], day_idx, schedule):
            continue
        result.append(shift)
    return result


def assign_shifts(staff_list, days_off_schedule, night_cover: dict = None):
    """
    Assigns shifts with the following priority order each day:
    Pass 1 — Fixed roles:
        a) night_fixed → Noche on working days
        b) designated normal employee → Noche when night_fixed is off
        c) jefe → Mañana on working days
    Pass 2 — Remaining workers: fill coverage deficits first, then preferences.
    """
    if night_cover is None:
        night_cover = {}

    schedule = copy.deepcopy(days_off_schedule)
    night_fixed = next((p for p in staff_list if p["role"] == "night_fixed"), None)

    # Pass 1: assign fixed roles
    for day_idx in range(len(DAYS)):
        for person in staff_list:
            if schedule[person["name"]][day_idx] == "Libre":
                continue
            if person["role"] == "night_fixed":
                schedule[person["name"]][day_idx] = "Noche"
            elif person["role"] == "jefe":
                schedule[person["name"]][day_idx] = "Mañana"
            elif person["name"] == night_cover.get(day_idx):
                # This normal employee covers Noche today (night_fixed is off)
                schedule[person["name"]][day_idx] = "Noche"

    # Pass 2: day by day, remaining workers
    for day_idx in range(len(DAYS)):
        available = [p for p in staff_list if schedule[p["name"]][day_idx] == ""]
        random.shuffle(available)

        # Current coverage
        coverage = {s: 0 for s in MAX_COVERAGE}
        for person in staff_list:
            s = schedule[person["name"]][day_idx]
            if s in coverage:
                coverage[s] += 1

        def unfilled_needs():
            needed = []
            for shift in ["Mañana", "Tarde", "Partido"]:
                deficit = MIN_COVERAGE.get(shift, 0) - coverage.get(shift, 0)
                if deficit > 0:
                    needed.extend([shift] * deficit)
            return needed

        # Sub-pass A: fill minimum coverage deficits
        for person in list(available):
            needed = unfilled_needs()
            if not needed:
                break
            opts = allowed_shifts(person, day_idx, schedule, coverage)
            fillable = [s for s in opts if s in needed]
            if not fillable:
                continue
            prefs = [s for s in person.get("preferences", []) if s in fillable]
            chosen = prefs[0] if prefs else fillable[0]
            schedule[person["name"]][day_idx] = chosen
            coverage[chosen] += 1
            available.remove(person)

        # Sub-pass B: assign remaining respecting caps
        for person in available:
            opts = allowed_shifts(person, day_idx, schedule, coverage)
            if not opts:
                fallback = [s for s in ["Partido", "Tarde", "Mañana"]
                            if coverage.get(s, 0) < MAX_COVERAGE.get(s, 99)
                            and (s != "Mañana" or can_work_morning(person["name"], day_idx, schedule))]
                chosen = fallback[0] if fallback else "Tarde"
            else:
                prefs = [s for s in person.get("preferences", []) if s in opts]
                chosen = min(prefs or opts, key=lambda s: coverage.get(s, 0))
            schedule[person["name"]][day_idx] = chosen
            coverage[chosen] = coverage.get(chosen, 0) + 1

    return schedule


def generate_schedule(staff_list=None, week_start: date = None):
    """Main function to generate a complete weekly schedule."""
    if staff_list is None:
        staff_list = copy.deepcopy(STAFF)
    if week_start is None:
        week_start = date.today() - timedelta(days=date.today().weekday())

    # select_night_covers also mutates chosen person's forced_days_off → Sáb+Dom
    night_cover = select_night_covers(staff_list, week_start)

    days_off_sched = assign_days_off(staff_list, night_cover)
    full_schedule = assign_shifts(staff_list, days_off_sched, night_cover)

    return full_schedule, staff_list, week_start


# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

SHIFT_COLORS = {
    "Mañana": "FFF59D",  # amarillo suave
    "Tarde": "A5D6A7",  # verde suave
    "Noche": "90CAF9",  # azul suave
    "Partido": "FFCC80",  # naranja suave
    "Libre": "EF9A9A",  # rojo suave
}

HEADER_COLOR = "1565C0"  # azul oscuro
SUB_HEADER = "1976D2"  # azul medio


def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def export_to_excel(schedule, staff_list, week_start, filename="horario_hotel.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario Semanal"

    # ── Title ──
    week_end = week_start + timedelta(days=6)
    title = f"Horario Semana  {week_start.strftime('%d/%m/%Y')} – {week_end.strftime('%d/%m/%Y')}"
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", start_color=HEADER_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Day headers (row 2) ──
    ws["A2"] = "Empleado"
    ws["A2"].font = Font(bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", start_color=SUB_HEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 18

    for col, (day, offset) in enumerate(zip(DAYS, range(7)), start=2):
        day_date = week_start + timedelta(days=offset)
        cell = ws.cell(row=2, column=col)
        cell.value = f"{day}\n{day_date.strftime('%d/%m')}"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", start_color=SUB_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.row_dimensions[2].height = 32

    # ── Data rows ──
    for row_idx, person in enumerate(staff_list, start=3):
        name_cell = ws.cell(row=row_idx, column=1, value=person["name"])
        role_label = ""
        if person["role"] == "night_fixed":
            role_label = " ★"
        elif person["role"] == "jefe":
            role_label = " 👑"
        elif person["role"] == "subjefe":
            role_label = " ⭐"
        elif person["role"] == "adjunto_recepcion":
            role_label = " ◉"
        name_cell.value = person["name"] + role_label
        name_cell.font = Font(bold=True)
        name_cell.alignment = Alignment(vertical="center", horizontal="left")
        name_cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 22

        for col_idx, day in enumerate(DAYS, start=2):
            shift = schedule[person["name"]][col_idx - 2]
            cell = ws.cell(row=row_idx, column=col_idx, value=shift)
            color = SHIFT_COLORS.get(shift, "FFFFFF")
            cell.fill = PatternFill("solid", start_color=color)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()
            cell.font = Font(size=10)

    # ── Coverage summary ──
    summary_row = len(staff_list) + 4
    ws.cell(row=summary_row, column=1, value="COBERTURA POR TURNO").font = Font(bold=True, size=11)

    shift_rows = {}
    for s_idx, shift in enumerate(["Mañana", "Tarde", "Noche", "Partido"]):
        r = summary_row + 1 + s_idx
        shift_rows[shift] = r
        label = ws.cell(row=r, column=1, value=shift)
        label.font = Font(bold=True)
        label.fill = PatternFill("solid", start_color=SHIFT_COLORS[shift])
        label.border = thin_border()
        label.alignment = Alignment(horizontal="center")

        for col_idx in range(2, 9):
            day_col = get_column_letter(col_idx)
            # Count using COUNTIF over the staff rows
            data_start = 3
            data_end = 3 + len(staff_list) - 1
            formula = f'=COUNTIF({day_col}{data_start}:{day_col}{data_end},"{shift}")'
            cell = ws.cell(row=r, column=col_idx, value=formula)
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border()
            cell.fill = PatternFill("solid", start_color=SHIFT_COLORS[shift])

    # ── Legend ──
    legend_row = summary_row + 7
    ws.cell(row=legend_row, column=1, value="LEYENDA").font = Font(bold=True)
    legends = [
        ("★ Turno Noche Fijo (libre Jue+Vie)", ""),
        ("👑 Jefe (Mañana fijo, libre Sáb+Dom)", ""),
        ("⭐ Subjefe", ""),
        ("◉ Adjunto Recepción (sin noches)", ""),
        ("Mañana", "FFF59D"),
        ("Tarde", "A5D6A7"),
        ("Noche", "90CAF9"),
        ("Partido", "FFCC80"),
        ("Libre", "EF9A9A"),
    ]
    for i, (label, color) in enumerate(legends):
        cell = ws.cell(row=legend_row + 1 + i, column=1, value=label)
        if color:
            cell.fill = PatternFill("solid", start_color=color)
        cell.border = thin_border()
        cell.alignment = Alignment(horizontal="left", vertical="center")

    # Freeze top rows
    ws.freeze_panes = "B3"

    wb.save(filename)
    return filename


# ─── DIAGNÓSTICO (uso interno, no ejecutar directamente) ─────────────────────

def diagnostico(staff_list=None, week_start: date = None):
    """
    Genera un horario de prueba e imprime el resumen de cobertura e infracciones.
    Útil para verificar que modelo.py funciona correctamente de forma aislada.
    Llama a esta función desde main.py o desde tests, no directamente.
    """
    import copy
    if staff_list is None:
        staff_list = copy.deepcopy(STAFF)
    schedule, staff, ws_date = generate_schedule(staff_list, week_start)

    print(f"\nSemana: {ws_date.strftime('%d/%m/%Y')} - {(ws_date + timedelta(days=6)).strftime('%d/%m/%Y')}\n")
    header = f"{'Empleado':<18}" + "".join(f"{d[:3]:<12}" for d in DAYS)
    print(header)
    print("-" * (18 + 12 * 7))
    for person in staff:
        row = f"{person['name']:<18}"
        for t in schedule[person["name"]]:
            row += f"{t:<12}"
        print(row)

    print("\n📊 Cobertura:")
    for shift in ["Mañana", "Tarde", "Noche", "Partido"]:
        for day_idx, day in enumerate(DAYS):
            count = sum(1 for p in staff if schedule[p["name"]][day_idx] == shift)
            mn, mx = MIN_COVERAGE.get(shift, 0), MAX_COVERAGE.get(shift, 99)
            ok = mn <= count <= mx
            if not ok:
                print(f"  ⚠ {day} {shift}: {count} (min {mn}, max {mx})")

    violations = [
        f"  ⚠ {p['name']}: {DAYS[i - 1]} ({schedule[p['name']][i - 1]}) → {DAYS[i]} (Mañana)"
        for p in staff
        for i in range(1, len(DAYS))
        if schedule[p["name"]][i - 1] in NO_MORNING_AFTER and schedule[p["name"]][i] == "Mañana"
    ]
    if violations:
        print("\n🔍 Infracciones Tarde/Noche → Mañana:")
        for v in violations:
            print(v)
    else:
        print("✅ Sin infracciones de turno.")

    return schedule, staff