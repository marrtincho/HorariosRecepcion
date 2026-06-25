"""
modelo_base.py — Motor genérico de generación de horarios parametrizable.
Recibe un DeptConfig con todas las reglas del departamento.
modelo.py y modelo_mozos.py importan de aquí.
"""

import random
import copy
from datetime import date, timedelta
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DAYS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
SHIFTS = ["Mañana", "Tarde", "Noche", "Partido", "Libre"]
NO_MORNING_AFTER = {"Noche", "Tarde"}

@dataclass
class DeptConfig:
    """Todas las reglas de un departamento."""
    dept_id: str                     # "recepcion" | "mozos"
    dept_label: str                  # "Recepción" | "Mozos"
    night_fixed_days_off: list       # índices de días libres del fijo de noche
    night_cover_compensation: list   # días libres que recibe el sustituto
    min_coverage: dict               # mínimo por turno por día
    max_coverage: dict               # máximo por turno por día
    no_night_roles: set              # roles que nunca hacen noche
    fixed_roles: dict                # role → turno fijo (ej: "jefe" → "Mañana")
    fixed_days_off: dict             # role → [idx_dias] fijos (ej: "jefe" → [5,6])
    cover_consecutive: bool          # ¿el sustituto necesita libres consecutivos?
    night_rotation_key: str          # clave en night_rotation.json
    allow_non_consecutive_cover: bool # mozos: el sustituto puede tener libres no consec
    prefer_next_week_for_cover: bool  # mozos: preferir Lun+Mar semana siguiente


# ─── UTILIDADES ─────────────────────────────────────────────────────────────

def get_consecutive_day_pairs():
    return [(i, i+1) for i in range(len(DAYS)-1)]

def _turno_anterior(person_name, day_idx, schedule, prev_sunday_shifts=None):
    """Turno del día anterior. Para el lunes (day_idx=0), mira el domingo
    de la semana pasada (prev_sunday_shifts), no reinicia la regla."""
    if day_idx == 0:
        return (prev_sunday_shifts or {}).get(person_name)
    return schedule[person_name][day_idx-1]

def can_work_morning(person_name, day_idx, schedule, prev_sunday_shifts=None):
    return _turno_anterior(person_name, day_idx, schedule, prev_sunday_shifts) not in NO_MORNING_AFTER

def _get_weekend_counts(dept_id: str) -> dict:
    try:
        from herramientas import cargar_historial, DATA_DIR
        import os, json
        historial = cargar_historial(dept_id)
        counts = {}
        for entrada in historial.values():
            for nombre, turnos in entrada.get("schedule", {}).items():
                if len(turnos) >= 7 and turnos[5] == "Libre" and turnos[6] == "Libre":
                    counts[nombre] = counts.get(nombre, 0) + 1
        return counts
    except Exception:
        return {}

def _load_night_history(rotation_key: str) -> dict:
    try:
        from herramientas import DATA_DIR
        import json, os
        path = os.path.join(DATA_DIR, f"night_rotation_{rotation_key}.json")
        return json.loads(open(path, encoding="utf-8").read()) if os.path.exists(path) else {}
    except Exception:
        return {}

def _save_night_history(rotation_key: str, history: dict):
    try:
        from herramientas import DATA_DIR
        import json, os
        path = os.path.join(DATA_DIR, f"night_rotation_{rotation_key}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ─── SELECCIÓN DEL SUSTITUTO NOCTURNO ───────────────────────────────────────

def select_night_covers(staff_list: list, week_start: date, cfg: DeptConfig) -> dict:
    """
    Selects one normal employee to cover the night_fixed's days off.
    - Recepción: covers Thu+Fri, gets Sat+Sun off.
    - Mozos: covers Fri+Sat nights, gets non-consecutive days off
      (preferably Mon+Tue of the NEXT week, handled via next_week_la flag).
    Returns {day_idx: chosen_name, ...}
    """
    blocked = set()
    for p in staff_list:
        forced = p.get("forced_days_off", [])
        for day_name in [DAYS[i] for i in cfg.night_fixed_days_off]:
            if day_name in forced:
                blocked.add(p["name"])

    try:
        from herramientas import cargar_roles
        roles_cfg = cargar_roles(cfg.dept_id)
        eligible_roles = {rid for rid, r in roles_cfg.items()
                           if not r.get("sin_noches") and rid != "night_fixed"}
    except Exception:
        eligible_roles = {"normal"}

    eligible = [
        p for p in staff_list
        if p["role"] in eligible_roles and p["name"] not in blocked
    ]
    if not eligible:
        return {}

    history = _load_night_history(cfg.night_rotation_key)

    try:
        from herramientas import cargar_compensaciones
        comp = cargar_compensaciones(cfg.dept_id)
    except Exception:
        comp = {}

    def sort_key(p):
        count = history.get(p["name"], 0)
        prefers_cobrar = 0 if comp.get(p["name"], {}).get("modalidad", "cobrar") == "cobrar" else 1
        return (count, prefers_cobrar, p["name"])

    chosen_person = sorted(eligible, key=sort_key)[0]
    chosen_name   = chosen_person["name"]

    # Compensation days off
    if cfg.prefer_next_week_for_cover:
        # Mark with a flag — actual days off assigned in next week's generation
        # For THIS week: give them the least-used consecutive pair that avoids cover days
        chosen_person["night_cover_this_week"] = True
        # They get no forced days off this week from coverage (their normal rotation applies)
        # but we note it so next week gives them Lun+Mar
        chosen_person["covered_nights_prev_week"] = True
        # Excepción mozos: si tras la última noche cubierta (p.ej. Sábado, si
        # cubre Vie+Sáb) todavía queda un día dentro de la MISMA semana antes
        # del domingo inclusive, ese día debe librarse obligatoriamente —
        # nunca puede empezar un turno normal recién saliendo de una noche.
        last_day = max(cfg.night_fixed_days_off) if cfg.night_fixed_days_off else -1
        if 0 <= last_day < 6:
            chosen_person["dia_descanso_tras_cobertura"] = last_day + 1
    else:
        # Recepción style: el sustituto descansa justo después de su última noche
        # de cobertura. Si esos días de compensación caen dentro de la misma
        # semana (p.ej. cubre Jue+Vie, descansa Sáb+Dom) se aplican ya — pero si
        # la última noche cubierta es Sábado o Domingo, "el día siguiente" cae
        # en la semana QUE VIENE (Domingo→Lunes), así que ese descanso no se
        # puede aplicar todavía: se guarda en next_week_rest_days para que el
        # llamador lo traslade a la generación de la semana siguiente.
        last_day = max(cfg.night_fixed_days_off) if cfg.night_fixed_days_off else -1
        same_week_idx = [i for i in cfg.night_cover_compensation if i > last_day]
        next_week_idx = [i for i in cfg.night_cover_compensation if i <= last_day]
        if same_week_idx:
            comp_day_names = [DAYS[i] for i in same_week_idx]
            existentes = chosen_person.get("forced_days_off", [])
            chosen_person["forced_days_off"] = list(dict.fromkeys(existentes + comp_day_names))
        if next_week_idx:
            chosen_person["next_week_rest_days"] = [DAYS[i] for i in next_week_idx]

    history[chosen_name] = history.get(chosen_name, 0) + 1
    _save_night_history(cfg.night_rotation_key, history)

    return {day_idx: chosen_name for day_idx in cfg.night_fixed_days_off}

# ─── ASIGNACIÓN DE LIBRES ────────────────────────────────────────────────────

def assign_days_off(staff_list: list, night_cover: dict, cfg: DeptConfig) -> dict:
    """Generic days-off assignment respecting dept config."""
    if night_cover is None:
        night_cover = {}

    cover_by_person = {}
    for day_idx, name in night_cover.items():
        cover_by_person.setdefault(name, []).append(day_idx)

    pairs          = get_consecutive_day_pairs()
    schedule       = {p["name"]: ["" for _ in DAYS] for p in staff_list}
    pair_usage     = {pair: 0 for pair in pairs}
    weekend_counts = _get_weekend_counts(cfg.dept_id)

    # Fixed roles: days off
    for role, days in cfg.fixed_days_off.items():
        person = next((p for p in staff_list if p["role"] == role), None)
        if person:
            for d in days:
                schedule[person["name"]][d] = "Libre"
            person["days_off"] = list(days)
            pair = tuple(sorted(days))
            if pair in pair_usage:
                pair_usage[pair] += 1

    # Night fixed: days off
    nf = next((p for p in staff_list if p["role"] == "night_fixed"), None)
    if nf:
        for d in cfg.night_fixed_days_off:
            schedule[nf["name"]][d] = "Libre"
        nf["days_off"] = list(cfg.night_fixed_days_off)
        pair = tuple(cfg.night_fixed_days_off)
        if pair in pair_usage:
            pair_usage[pair] += 1

    # Vacaciones y peticiones de libre: se respetan siempre y por completo,
    # sin importar el rol (incluso para turnos fijos como jefe o night_fixed).
    for person in staff_list:
        forced_indices = [DAYS.index(d) for d in person.get("forced_days_off", []) if d in DAYS]
        extra_indices  = [DAYS.index(d) for d in person.get("extra_days_off", [])  if d in DAYS]
        for idx in forced_indices + extra_indices:
            schedule[person["name"]][idx] = "Libre"
        if forced_indices or extra_indices:
            person["days_off"] = sorted(set(person.get("days_off", [])) | set(forced_indices) | set(extra_indices))

    # Everyone else (roles sin turno fijo): completar su pareja de libres semanales
    fixed_roles = set(cfg.fixed_days_off.keys()) | {"night_fixed"}
    for person in staff_list:
        if person["role"] in fixed_roles:
            continue

        forced  = person.get("forced_days_off", [])
        forced_indices = [DAYS.index(d) for d in forced if d in DAYS]
        must_work = set(cover_by_person.get(person["name"], []))

        if len(forced_indices) >= 2:
            # Ya quedaron marcados Libre arriba; solo se registra el par
            # principal para el balanceo de uso de parejas de días.
            pair = tuple(sorted(forced_indices[:2]))
            if pair in pair_usage:
                pair_usage[pair] += 1

        elif len(forced_indices) == 1:
            fi = forced_indices[0]
            candidates = [(fi-1,fi)] if fi>0 else []
            if fi < len(DAYS)-1:
                candidates.append((fi, fi+1))
            valid = [c for c in candidates if c in pair_usage and not (must_work & set(c))]
            if not valid:
                valid = [c for c in candidates if c in pair_usage]
            chosen = min(valid, key=lambda p: pair_usage[p]) if valid \
                     else sorted(pairs, key=lambda p: pair_usage[p])[0]
            schedule[person["name"]][chosen[0]] = "Libre"
            schedule[person["name"]][chosen[1]] = "Libre"
            person["days_off"] = list(chosen)
            pair_usage[chosen] += 1

        elif person.get("dia_descanso_tras_cobertura") is not None:
            # Mozos cover (Vie+Sáb): si tras la última noche cubierta aún
            # queda un día dentro de la misma semana (el domingo), ese día
            # libra obligatoriamente como descanso — sin pareja de rotación
            # normal aparte, para no superar los 2 libres semanales (el
            # premio de Lun+Mar de la semana siguiente se aplica por separado).
            dia = person["dia_descanso_tras_cobertura"]
            schedule[person["name"]][dia] = "Libre"
            person["days_off"] = sorted(set(person.get("days_off", [])) | {dia})

        elif person.get("night_cover_this_week"):
            # Mozos cover: no forced comp days this week, use normal rotation
            # but avoid covering nights (those are their work days)
            valid_pairs = [p for p in pairs if not (must_work & set(p))]
            if not valid_pairs:
                valid_pairs = pairs
            chosen = sorted(valid_pairs, key=lambda p: pair_usage[p])[0]
            pair_usage[chosen] += 1
            schedule[person["name"]][chosen[0]] = "Libre"
            schedule[person["name"]][chosen[1]] = "Libre"
            person["days_off"] = list(chosen)

        else:
            # Normal rotation with weekend balancing
            valid_pairs = [p for p in pairs if not (must_work & set(p))]
            if not valid_pairs:
                valid_pairs = pairs
            wc = weekend_counts.get(person["name"], 0)
            def pair_sort_key(pair):
                usage = pair_usage[pair]
                weekend_bonus = -0.5 if pair == (5,6) and wc < 1 else 0
                return usage + weekend_bonus
            chosen = sorted(valid_pairs, key=pair_sort_key)[0]
            pair_usage[chosen] += 1
            schedule[person["name"]][chosen[0]] = "Libre"
            schedule[person["name"]][chosen[1]] = "Libre"
            person["days_off"] = list(chosen)
            if chosen == (5,6):
                weekend_counts[person["name"]] = wc + 1

    return schedule

# ─── ASIGNACIÓN DE TURNOS ────────────────────────────────────────────────────

def allowed_shifts(person, day_idx, schedule, coverage, cfg: DeptConfig, prev_sunday_shifts=None):
    result = []
    for shift in ["Mañana", "Tarde", "Partido"]:
        if coverage.get(shift, 0) >= cfg.max_coverage.get(shift, 99):
            continue
        if shift == "Mañana" and not can_work_morning(person["name"], day_idx, schedule, prev_sunday_shifts):
            continue
        result.append(shift)
    return result

def assign_shifts(staff_list: list, days_off_schedule: dict,
                  night_cover: dict, cfg: DeptConfig, prev_sunday_shifts: dict = None) -> dict:
    if night_cover is None:
        night_cover = {}

    schedule = copy.deepcopy(days_off_schedule)

    # Pass 1: fixed roles
    for day_idx in range(len(DAYS)):
        for person in staff_list:
            if schedule[person["name"]][day_idx] == "Libre":
                continue
            if person["role"] == "night_fixed":
                schedule[person["name"]][day_idx] = "Noche"
            elif person["name"] == night_cover.get(day_idx):
                schedule[person["name"]][day_idx] = "Noche"
            elif person["role"] in cfg.fixed_roles:
                schedule[person["name"]][day_idx] = cfg.fixed_roles[person["role"]]

    # Pass 2: day by day
    for day_idx in range(len(DAYS)):
        available = [p for p in staff_list if schedule[p["name"]][day_idx] == ""]
        random.shuffle(available)

        coverage = {s: 0 for s in cfg.max_coverage}
        for person in staff_list:
            s = schedule[person["name"]][day_idx]
            if s in coverage:
                coverage[s] += 1

        def unfilled():
            needed = []
            for shift in ["Mañana", "Tarde", "Partido"]:
                deficit = cfg.min_coverage.get(shift, 0) - coverage.get(shift, 0)
                if deficit > 0:
                    needed.extend([shift] * deficit)
            return needed

        # Sub-pass A: fill deficits
        for person in list(available):
            needed = unfilled()
            if not needed:
                break
            opts = allowed_shifts(person, day_idx, schedule, coverage, cfg, prev_sunday_shifts)
            fillable = [s for s in opts if s in needed]
            if not fillable:
                continue
            prefs = [s for s in person.get("preferences", []) if s in fillable]
            chosen = prefs[0] if prefs else fillable[0]
            schedule[person["name"]][day_idx] = chosen
            coverage[chosen] += 1
            available.remove(person)

        # Sub-pass B: remaining
        for person in available:
            opts = allowed_shifts(person, day_idx, schedule, coverage, cfg, prev_sunday_shifts)
            if not opts:
                fallback = [s for s in ["Partido", "Tarde", "Mañana"]
                            if coverage.get(s, 0) < cfg.max_coverage.get(s, 99)
                            and (s != "Mañana" or can_work_morning(person["name"], day_idx, schedule, prev_sunday_shifts))]
                chosen = fallback[0] if fallback else "Tarde"
            else:
                prefs = [s for s in person.get("preferences", []) if s in opts]
                chosen = min(prefs or opts, key=lambda s: coverage.get(s, 0))
            schedule[person["name"]][day_idx] = chosen
            coverage[chosen] = coverage.get(chosen, 0) + 1

    return schedule

# ─── FUNCIÓN PRINCIPAL ───────────────────────────────────────────────────────

def generate_schedule(staff_list: list, week_start: date, cfg: DeptConfig,
                      prev_week_covers: dict = None, prev_sunday_shifts: dict = None):
    """
    Main schedule generation function.
    prev_week_covers: {name: True} (mozos) o {name: [día,...]} (recepción) —
    quién cubrió noches la semana pasada y debe descansar esos días esta semana,
    porque su descanso de compensación cayó tras el fin de la semana anterior.
    prev_sunday_shifts: {name: turno} — turno del domingo de la semana anterior,
    para que las reglas de descanso (no Mañana tras Tarde/Noche, descanso tras Noche)
    no se reinicien al cambiar de semana.
    """
    staff_list = copy.deepcopy(staff_list)

    # Aplicar el descanso de compensación pendiente de la semana pasada.
    if prev_week_covers:
        for person in staff_list:
            dias = prev_week_covers.get(person["name"])
            if not dias:
                continue
            if cfg.prefer_next_week_for_cover:
                # Mozos: siempre Lun+Mar de la semana siguiente, por diseño.
                person["forced_days_off"] = ["Lunes", "Martes"]
            elif isinstance(dias, list):
                existentes = person.get("forced_days_off", [])
                person["forced_days_off"] = list(dict.fromkeys(existentes + dias))

    night_cover    = select_night_covers(staff_list, week_start, cfg)
    days_off_sched = assign_days_off(staff_list, night_cover, cfg)
    full_schedule  = assign_shifts(staff_list, days_off_sched, night_cover, cfg, prev_sunday_shifts)

    # Registrar quién necesita descanso de compensación la semana que viene.
    this_week_covers = {}
    if cfg.prefer_next_week_for_cover:
        for person in staff_list:
            if person.get("night_cover_this_week"):
                this_week_covers[person["name"]] = True
    else:
        for person in staff_list:
            dias = person.get("next_week_rest_days")
            if dias:
                this_week_covers[person["name"]] = dias

    return full_schedule, staff_list, week_start, night_cover, this_week_covers

# ─── EXCEL EXPORT ────────────────────────────────────────────────────────────

SHIFT_COLORS = {
    "Mañana":  "FFF59D",
    "Tarde":   "A5D6A7",
    "Noche":   "90CAF9",
    "Partido": "FFCC80",
    "Libre":   "EF9A9A",
}
HEADER_COLOR = "1565C0"
SUB_HEADER   = "1976D2"

def thin_border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def export_to_excel(schedule, staff_list, week_start, filename, cfg: DeptConfig, role_icons: dict = None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Horario Semanal"

    week_end = week_start + timedelta(days=6)
    title = f"{cfg.dept_label} — Semana {week_start.strftime('%d/%m/%Y')} – {week_end.strftime('%d/%m/%Y')}"
    ws.merge_cells("A1:I1")
    ws["A1"] = title
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color=HEADER_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "Empleado"
    ws["A2"].font      = Font(bold=True, color="FFFFFF")
    ws["A2"].fill      = PatternFill("solid", start_color=SUB_HEADER)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 18

    for col, offset in enumerate(range(7), start=2):
        day_date = week_start + timedelta(days=offset)
        cell = ws.cell(row=2, column=col)
        cell.value = f"{DAYS[offset]}\n{day_date.strftime('%d/%m')}"
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=SUB_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 13
    ws.row_dimensions[2].height = 32

    icons = role_icons if role_icons is not None else \
        {"night_fixed":"★","jefe":"👑","subjefe":"⭐","adjunto_recepcion":"◉"}
    for row_idx, person in enumerate(staff_list, start=3):
        icon = icons.get(person["role"], "")
        name_cell = ws.cell(row=row_idx, column=1, value=f"{person['name']} {icon}".strip())
        name_cell.font      = Font(bold=True)
        name_cell.alignment = Alignment(vertical="center", horizontal="left")
        name_cell.border    = thin_border()
        ws.row_dimensions[row_idx].height = 22
        for col_idx in range(7):
            shift = schedule[person["name"]][col_idx]
            cell  = ws.cell(row=row_idx, column=col_idx+2, value=shift)
            cell.fill      = PatternFill("solid", start_color=SHIFT_COLORS.get(shift,"FFFFFF"))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()
            cell.font      = Font(size=10)

    summary_row = len(staff_list) + 4
    ws.cell(row=summary_row, column=1, value="COBERTURA").font = Font(bold=True, size=11)
    for s_idx, shift in enumerate(["Mañana","Tarde","Noche","Partido"]):
        r = summary_row + 1 + s_idx
        label = ws.cell(row=r, column=1, value=shift)
        label.font = Font(bold=True)
        label.fill = PatternFill("solid", start_color=SHIFT_COLORS[shift])
        label.border = thin_border()
        label.alignment = Alignment(horizontal="center")
        for col_idx in range(2,9):
            day_col = get_column_letter(col_idx)
            ds, de  = 3, 3 + len(staff_list) - 1
            cell = ws.cell(row=r, column=col_idx,
                           value=f'=COUNTIF({day_col}{ds}:{day_col}{de},"{shift}")')
            cell.alignment = Alignment(horizontal="center")
            cell.border    = thin_border()
            cell.fill      = PatternFill("solid", start_color=SHIFT_COLORS[shift])

    ws.freeze_panes = "B3"
    wb.save(filename)
    return filename
