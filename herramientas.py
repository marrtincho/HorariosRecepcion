"""
herramientas.py — Gestión de plantilla, peticiones semanales e historial.
Soporte multi-departamento: todas las funciones clave aceptan dept_id.
"""

import json
import os
import shutil
from datetime import date, timedelta

# ─── RUTAS DE DATOS ──────────────────────────────────────────────────────────

BASE_DIR        = os.path.dirname(os.path.abspath(__file__))
DATA_DIR        = os.path.join(BASE_DIR, "data")
BACKUP_DIR      = os.path.join(DATA_DIR, "backups")

os.makedirs(DATA_DIR,   exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

# Rutas por defecto (recepcion, retrocompatibilidad)
PLANTILLA_FILE  = os.path.join(DATA_DIR, "plantilla.json")
PETICIONES_FILE = os.path.join(DATA_DIR, "peticiones.json")
HISTORIAL_FILE  = os.path.join(DATA_DIR, "historial.json")
VACACIONES_FILE = os.path.join(DATA_DIR, "vacaciones.json")

DEPTS_VALIDOS = {"recepcion", "mozos"}

def _dept_file(base_name: str, dept_id: str = "recepcion") -> str:
    """Returns dept-namespaced file path. recepcion uses legacy names for compatibility."""
    if dept_id == "recepcion":
        return os.path.join(DATA_DIR, base_name)
    return os.path.join(DATA_DIR, f"{dept_id}_{base_name}")

# Roles válidos por departamento
ROLES_VALIDOS = {
    "recepcion": {
        "jefe":              "Jefe (Mañana fijo, libre Sáb+Dom)",
        "subjefe":           "Subjefe (preferencia Tarde)",
        "night_fixed":       "Turno Noche fijo (libre Jue+Vie)",
        "adjunto_recepcion": "Adjunto Recepción (sin noches)",
        "normal":            "Empleado normal",
    },
    "mozos": {
        "night_fixed": "Turno Noche fijo (libre Vie+Sáb)",
        "normal":      "Mozo normal",
    },
}

TURNOS_VALIDOS = ["Mañana", "Tarde", "Noche", "Partido"]
DIAS_SEMANA    = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]

# ─── UTILIDADES INTERNAS ─────────────────────────────────────────────────────

def _cargar_json(path, default):
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return default

def _guardar_json(path, data):
    """Guarda JSON haciendo backup automático del archivo anterior."""
    if os.path.exists(path):
        nombre = os.path.basename(path)
        ts     = date.today().strftime("%Y%m%d")
        backup = os.path.join(BACKUP_DIR, f"{nombre}.{ts}.bak")
        shutil.copy2(path, backup)
        # Mantener solo los últimos 10 backups por archivo para no acumular
        todos = sorted(
            f for f in os.listdir(BACKUP_DIR) if f.startswith(nombre)
        )
        for viejo in todos[:-10]:
            os.remove(os.path.join(BACKUP_DIR, viejo))
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def _limpiar_pantalla():
    os.system("cls" if os.name == "nt" else "clear")

def _pausar():
    input("\n  Pulsa Enter para continuar...")

def _separador(titulo=""):
    ancho = 55
    print("\n" + "─" * ancho)
    if titulo:
        print(f"  {titulo}")
        print("─" * ancho)

def _input_opcion(opciones: list[str], prompt="  Opción: ") -> str:
    """Pide una opción hasta que sea válida."""
    while True:
        v = input(prompt).strip()
        if v in opciones:
            return v
        print(f"  ⚠ Opción no válida. Elige entre: {', '.join(opciones)}")

def _semana_str(week_start: date) -> str:
    return week_start.strftime("%Y-%m-%d")

def _semana_desde_str(s: str) -> date:
    return date.fromisoformat(s)

# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 1 — PLANTILLA
# ═══════════════════════════════════════════════════════════════════════════════

# Plantillas por defecto por departamento
PLANTILLA_DEFAULT = {
    "recepcion": [
        {"name": "Carlos",  "role": "night_fixed",       "preferences": ["Noche"]},
        {"name": "Juan",    "role": "jefe",               "preferences": ["Mañana"]},
        {"name": "Laura",   "role": "subjefe",            "preferences": ["Tarde"]},
        {"name": "Pedro",   "role": "normal",             "preferences": ["Mañana", "Partido"]},
        {"name": "Ana",     "role": "normal",             "preferences": []},
        {"name": "Miguel",  "role": "normal",             "preferences": ["Tarde"]},
        {"name": "Sofía",   "role": "adjunto_recepcion",  "preferences": ["Mañana"]},
        {"name": "Diego",   "role": "normal",             "preferences": []},
    ],
    "mozos": [
        {"name": "Fran",   "role": "night_fixed", "preferences": ["Noche"]},
        {"name": "Rubén",  "role": "normal",      "preferences": ["Mañana"]},
        {"name": "Gloria", "role": "normal",      "preferences": ["Tarde"]},
        {"name": "Tomás",  "role": "normal",      "preferences": []},
        {"name": "Inés",   "role": "normal",      "preferences": ["Mañana"]},
    ],
}

# Roles únicos por departamento
ROLES_UNICOS = {
    "recepcion": {"jefe", "night_fixed"},
    "mozos":     {"night_fixed"},
}

def cargar_plantilla(dept_id: str = "recepcion") -> list[dict]:
    import copy
    path  = _dept_file("plantilla.json", dept_id)
    datos = _cargar_json(path, None)
    if datos is not None:
        return datos
    plantilla = copy.deepcopy(PLANTILLA_DEFAULT.get(dept_id, []))
    guardar_plantilla(plantilla, dept_id)
    print(f"  ℹ Plantilla {dept_id} inicializada con datos por defecto.")
    return plantilla

def guardar_plantilla(staff: list[dict], dept_id: str = "recepcion"):
    _guardar_json(_dept_file("plantilla.json", dept_id), staff)

def validar_plantilla(staff: list[dict], excluir_idx: int = None,
                      dept_id: str = "recepcion") -> list[str]:
    """Comprueba restricciones de integridad. Devuelve lista de errores (vacía = OK)."""
    errores = []
    unicos  = ROLES_UNICOS.get(dept_id, set())
    conteo_roles = {}
    for i, p in enumerate(staff):
        if i == excluir_idx:
            continue
        rol = p["role"]
        conteo_roles[rol] = conteo_roles.get(rol, 0) + 1

    roles_label = ROLES_VALIDOS.get(dept_id, {})
    for rol in unicos:
        if conteo_roles.get(rol, 0) > 1:
            label = roles_label.get(rol, rol)
            errores.append(f"El rol '{label}' solo puede asignarse a una persona (hay {conteo_roles[rol]}).")

    return errores

def _mostrar_plantilla(staff: list[dict]):
    _separador("PLANTILLA ACTUAL")
    print(f"  {'#':<4} {'Nombre':<16} {'Rol':<34} {'Preferencias'}")
    print("  " + "-" * 70)
    for i, p in enumerate(staff, 1):
        prefs = ", ".join(p.get("preferences", [])) or "—"
        rol   = ROLES_VALIDOS.get(p["role"], p["role"])
        print(f"  {i:<4} {p['name']:<16} {rol:<34} {prefs}")

def menu_plantilla():
    """Submenú completo de gestión de plantilla."""
    while True:
        staff = cargar_plantilla()
        _limpiar_pantalla()
        _mostrar_plantilla(staff)
        _separador("OPCIONES")
        print("  1. Añadir empleado")
        print("  2. Editar empleado")
        print("  3. Eliminar empleado")
        print("  0. Volver")
        opcion = _input_opcion(["1", "2", "3", "0"])

        if opcion == "1":
            _añadir_empleado(staff)
        elif opcion == "2":
            _editar_empleado(staff)
        elif opcion == "3":
            _eliminar_empleado(staff)
        elif opcion == "0":
            break

def _pedir_nombre(staff, excluir_idx=None) -> str:
    nombres = {p["name"].lower() for i, p in enumerate(staff) if i != excluir_idx}
    while True:
        nombre = input("  Nombre: ").strip()
        if not nombre:
            print("  ⚠ El nombre no puede estar vacío.")
        elif nombre.lower() in nombres:
            print("  ⚠ Ya existe un empleado con ese nombre.")
        else:
            return nombre

def _pedir_rol(staff: list[dict], excluir_idx: int = None) -> str:
    """Muestra roles disponibles e indica cuáles ya están ocupados (roles únicos)."""
    ocupados = {p["role"] for i, p in enumerate(staff) if i != excluir_idx}
    print("\n  Roles disponibles:")
    claves = list(ROLES_VALIDOS.keys())
    for i, (k, v) in enumerate(ROLES_VALIDOS.items(), 1):
        nota = "  ⚠ ya asignado — solo puede haber uno" if k in ROLES_UNICOS and k in ocupados else ""
        print(f"    {i}. {v}{nota}")
    opciones = [str(i) for i in range(1, len(claves) + 1)]
    while True:
        idx = int(_input_opcion(opciones, "  Rol (número): ")) - 1
        rol_elegido = claves[idx]
        if rol_elegido in ROLES_UNICOS and rol_elegido in ocupados:
            print(f"  ⚠ '{ROLES_VALIDOS[rol_elegido]}' ya está asignado a otra persona. Elige otro rol.")
        else:
            return rol_elegido

def _pedir_preferencias() -> list[str]:
    print(f"\n  Turnos disponibles: {', '.join(TURNOS_VALIDOS)}")
    print("  Escribe las preferencias en orden separadas por coma (o Enter para ninguna):")
    raw = input("  Preferencias: ").strip()
    if not raw:
        return []
    prefs = [p.strip().capitalize() for p in raw.split(",")]
    validas = [p for p in prefs if p in TURNOS_VALIDOS]
    ignoradas = [p for p in prefs if p not in TURNOS_VALIDOS]
    if ignoradas:
        print(f"  ⚠ Ignoradas (no válidas): {', '.join(ignoradas)}")
    return validas

def _añadir_empleado(staff: list[dict]):
    _separador("AÑADIR EMPLEADO")
    nombre = _pedir_nombre(staff)
    rol    = _pedir_rol(staff)
    prefs  = _pedir_preferencias()
    nuevo  = {"name": nombre, "role": rol, "preferences": prefs}
    staff.append(nuevo)
    errores = validar_plantilla(staff)
    if errores:
        staff.pop()
        print("\n  ❌ No se pudo añadir el empleado:")
        for e in errores:
            print(f"     • {e}")
    else:
        guardar_plantilla(staff)
        print(f"\n  ✅ Empleado '{nombre}' añadido correctamente.")
    _pausar()

def _editar_empleado(staff: list[dict]):
    _separador("EDITAR EMPLEADO")
    if not staff:
        print("  No hay empleados.")
        _pausar()
        return
    _mostrar_plantilla(staff)
    idx_str = _input_opcion([str(i) for i in range(1, len(staff) + 1)], "\n  Número de empleado a editar: ")
    idx = int(idx_str) - 1
    import copy
    p_original = copy.deepcopy(staff[idx])
    p = staff[idx]

    print(f"\n  Editando: {p['name']} (Enter para mantener valor actual)")

    nuevo_nombre = input(f"  Nombre [{p['name']}]: ").strip()
    if nuevo_nombre:
        # Check name collision excluding self
        otros = {s["name"].lower() for i, s in enumerate(staff) if i != idx}
        if nuevo_nombre.lower() in otros:
            print("  ⚠ Ese nombre ya existe. Se mantiene el nombre actual.")
        else:
            p["name"] = nuevo_nombre

    print(f"  Rol actual: {ROLES_VALIDOS.get(p['role'], p['role'])}")
    cambiar_rol = input("  ¿Cambiar rol? (s/n): ").strip().lower()
    if cambiar_rol == "s":
        p["role"] = _pedir_rol(staff, excluir_idx=idx)

    print(f"  Preferencias actuales: {', '.join(p.get('preferences', [])) or '—'}")
    cambiar_prefs = input("  ¿Cambiar preferencias? (s/n): ").strip().lower()
    if cambiar_prefs == "s":
        p["preferences"] = _pedir_preferencias()

    errores = validar_plantilla(staff, excluir_idx=idx)
    if errores:
        staff[idx] = p_original  # rollback
        print("\n  ❌ Cambios revertidos por errores de validación:")
        for e in errores:
            print(f"     • {e}")
    else:
        guardar_plantilla(staff)
        print(f"\n  ✅ Empleado actualizado correctamente.")
    _pausar()

def _eliminar_empleado(staff: list[dict]):
    _separador("ELIMINAR EMPLEADO")
    if not staff:
        print("  No hay empleados.")
        _pausar()
        return
    _mostrar_plantilla(staff)
    idx_str = _input_opcion([str(i) for i in range(1, len(staff) + 1)], "\n  Número de empleado a eliminar: ")
    idx    = int(idx_str) - 1
    nombre = staff[idx]["name"]
    rol    = staff[idx]["role"]
    if rol in ROLES_UNICOS:
        print(f"\n  ⚠ Atención: '{nombre}' tiene el rol único '{ROLES_VALIDOS.get(rol, rol)}'.")
        print("     Deberías asignar ese rol a otra persona antes de eliminar.")
    confirmar = input(f"  ¿Confirmar eliminación de '{nombre}'? (s/n): ").strip().lower()
    if confirmar == "s":
        staff.pop(idx)
        guardar_plantilla(staff)
        print(f"\n  ✅ Empleado '{nombre}' eliminado.")
    else:
        print("  Cancelado.")
    _pausar()


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 2 — PETICIONES SEMANALES
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_peticiones(dept_id: str = "recepcion") -> dict:
    return _cargar_json(_dept_file("peticiones.json", dept_id), {})

def guardar_peticiones(peticiones: dict, dept_id: str = "recepcion"):
    _guardar_json(_dept_file("peticiones.json", dept_id), peticiones)

def obtener_peticiones_semana(week_start: date, dept_id: str = "recepcion") -> dict:
    peticiones = cargar_peticiones(dept_id)
    return peticiones.get(_semana_str(week_start), {})

def aplicar_peticiones(staff_list: list[dict], week_start: date, dept_id: str = "recepcion") -> list[dict]:
    """
    Devuelve una copia de staff_list con:
    - 'forced_days_off': días de libre solicitados normalmente (peticiones regulares)
    - 'extra_days_off':  días LA (libres extra que se añaden al par normal)
    Soporta tanto el formato antiguo (lista) como el nuevo (dict con 'dias' y 'la').
    """
    import copy
    peticiones = obtener_peticiones_semana(week_start, dept_id)
    staff_copia = copy.deepcopy(staff_list)
    for person in staff_copia:
        entrada = peticiones.get(person["name"])
        if not entrada:
            continue
        if isinstance(entrada, list):
            # Formato antiguo — solo días regulares
            if entrada:
                person["forced_days_off"] = entrada
        elif isinstance(entrada, dict):
            dias_regulares = entrada.get("dias", [])
            dias_la        = entrada.get("la",   [])
            if dias_regulares:
                person["forced_days_off"] = dias_regulares
            if dias_la:
                person["extra_days_off"] = dias_la
    return staff_copia

def menu_peticiones():
    """Submenú de gestión de peticiones semanales."""
    while True:
        _limpiar_pantalla()
        _separador("PETICIONES SEMANALES")
        print("  1. Ver peticiones de una semana")
        print("  2. Añadir petición de día libre")
        print("  3. Eliminar petición")
        print("  0. Volver")
        opcion = _input_opcion(["1", "2", "3", "0"])

        if opcion == "1":
            _ver_peticiones()
        elif opcion == "2":
            _añadir_peticion()
        elif opcion == "3":
            _eliminar_peticion()
        elif opcion == "0":
            break

def _pedir_semana(prompt="  Semana (YYYY-MM-DD del lunes, Enter = semana actual): ") -> date:
    while True:
        raw = input(prompt).strip()
        if not raw:
            hoy = date.today()
            return hoy - timedelta(days=hoy.weekday())
        try:
            d = date.fromisoformat(raw)
            # Ajustar al lunes de esa semana
            return d - timedelta(days=d.weekday())
        except ValueError:
            print("  ⚠ Formato incorrecto. Usa YYYY-MM-DD.")

def _ver_peticiones():
    _separador("VER PETICIONES")
    semana = _pedir_semana()
    peticiones = obtener_peticiones_semana(semana)
    print(f"\n  Semana del {semana.strftime('%d/%m/%Y')}:")
    if not peticiones:
        print("  Sin peticiones registradas.")
    else:
        for nombre, dias in peticiones.items():
            print(f"    {nombre}: {', '.join(dias)}")
    _pausar()

def _añadir_peticion():
    _separador("AÑADIR PETICIÓN")
    staff   = cargar_plantilla()
    semana  = _pedir_semana()
    semana_str = _semana_str(semana)

    print("\n  Empleados disponibles:")
    for i, p in enumerate(staff, 1):
        print(f"    {i}. {p['name']}")
    idx_str = _input_opcion([str(i) for i in range(1, len(staff) + 1)], "  Número de empleado: ")
    nombre  = staff[int(idx_str) - 1]["name"]

    print(f"\n  Días: {', '.join(DIAS_SEMANA)}")
    print("  Escribe los días libres solicitados separados por coma:")
    raw  = input("  Días: ").strip()
    dias = [d.strip().capitalize() for d in raw.split(",")]
    # Normalizar miércoles
    dias = [d if d != "Miercoles" else "Miércoles" for d in dias]
    validos  = [d for d in dias if d in DIAS_SEMANA]
    invalidos = [d for d in dias if d not in DIAS_SEMANA]
    if invalidos:
        print(f"  ⚠ Días ignorados (no válidos): {', '.join(invalidos)}")
    if not validos:
        print("  No se registró ningún día válido.")
        _pausar()
        return

    peticiones = cargar_peticiones()
    if semana_str not in peticiones:
        peticiones[semana_str] = {}
    peticiones[semana_str][nombre] = validos
    guardar_peticiones(peticiones)
    print(f"\n  ✅ Petición guardada: {nombre} → {', '.join(validos)} (semana {semana.strftime('%d/%m/%Y')})")
    _pausar()

def _eliminar_peticion():
    _separador("ELIMINAR PETICIÓN")
    semana     = _pedir_semana()
    semana_str = _semana_str(semana)
    peticiones = cargar_peticiones()
    semana_pet = peticiones.get(semana_str, {})

    if not semana_pet:
        print(f"  No hay peticiones para la semana del {semana.strftime('%d/%m/%Y')}.")
        _pausar()
        return

    print(f"\n  Peticiones semana {semana.strftime('%d/%m/%Y')}:")
    nombres = list(semana_pet.keys())
    for i, n in enumerate(nombres, 1):
        print(f"    {i}. {n}: {', '.join(semana_pet[n])}")

    idx_str = _input_opcion([str(i) for i in range(1, len(nombres) + 1)], "  Número a eliminar: ")
    nombre  = nombres[int(idx_str) - 1]
    del peticiones[semana_str][nombre]
    if not peticiones[semana_str]:
        del peticiones[semana_str]
    guardar_peticiones(peticiones)
    print(f"\n  ✅ Petición de '{nombre}' eliminada.")
    _pausar()


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 3 — HISTORIAL
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_historial(dept_id: str = "recepcion") -> dict:
    return _cargar_json(_dept_file("historial.json", dept_id), {})

def guardar_historial(historial: dict, dept_id: str = "recepcion"):
    _guardar_json(_dept_file("historial.json", dept_id), historial)

def registrar_semana(week_start: date, schedule: dict, staff_list: list[dict],
                     forzar: bool = False, dept_id: str = "recepcion") -> bool:
    historial  = cargar_historial(dept_id)
    semana_str = _semana_str(week_start)
    if semana_str in historial and not forzar:
        return False
    historial[semana_str] = {
        "schedule": schedule,
        "staff":    [{k: v for k, v in p.items() if k != "days_off"} for p in staff_list],
    }
    guardar_historial(historial, dept_id)
    return True

def menu_historial():
    """Submenú de consulta de historial y estadísticas."""
    while True:
        _limpiar_pantalla()
        _separador("HISTORIAL DE SEMANAS")
        historial = cargar_historial()
        semanas   = sorted(historial.keys(), reverse=True)

        if not semanas:
            print("  No hay semanas registradas aún.")
            _pausar()
            return

        print(f"  Semanas guardadas: {len(semanas)}")
        print("  1. Ver horario de una semana")
        print("  2. Estadísticas por empleado")
        print("  3. Eliminar semana del historial")
        print("  0. Volver")
        opcion = _input_opcion(["1", "2", "3", "0"])

        if opcion == "1":
            _ver_semana_historial(historial, semanas)
        elif opcion == "2":
            _estadisticas(historial)
        elif opcion == "3":
            _eliminar_semana_historial(historial, semanas)
        elif opcion == "0":
            break

def _seleccionar_semana_historial(semanas: list[str]) -> str | None:
    print("\n  Semanas disponibles:")
    for i, s in enumerate(semanas, 1):
        d = _semana_desde_str(s)
        fin = d + timedelta(days=6)
        print(f"    {i}. {d.strftime('%d/%m/%Y')} – {fin.strftime('%d/%m/%Y')}")
    opciones = [str(i) for i in range(1, len(semanas) + 1)]
    idx_str  = _input_opcion(opciones, "  Número de semana: ")
    return semanas[int(idx_str) - 1]

def _ver_semana_historial(historial: dict, semanas: list[str]):
    _separador("VER SEMANA")
    key      = _seleccionar_semana_historial(semanas)
    entrada  = historial[key]
    schedule = entrada["schedule"]
    staff    = entrada["staff"]
    d_ini    = _semana_desde_str(key)

    print(f"\n  Semana del {d_ini.strftime('%d/%m/%Y')}:\n")
    header = f"  {'Empleado':<16}" + "".join(f"{d[:3]:<10}" for d in DIAS_SEMANA)
    print(header)
    print("  " + "-" * (16 + 10 * 7))
    for p in staff:
        fila = f"  {p['name']:<16}"
        for t in schedule.get(p["name"], ["?" ] * 7):
            fila += f"{t:<10}"
        print(fila)
    _pausar()

def _estadisticas(historial: dict):
    _separador("ESTADÍSTICAS POR EMPLEADO")

    conteos: dict[str, dict[str, int]] = {}
    semanas_trabajadas: dict[str, int] = {}

    for semana_str, entrada in historial.items():
        schedule = entrada["schedule"]
        for nombre, turnos in schedule.items():
            if nombre not in conteos:
                conteos[nombre] = {t: 0 for t in ["Mañana", "Tarde", "Noche", "Partido", "Libre"]}
                semanas_trabajadas[nombre] = 0
            semanas_trabajadas[nombre] += 1
            for t in turnos:
                if t in conteos[nombre]:
                    conteos[nombre][t] += 1

    if not conteos:
        print("  Sin datos.")
        _pausar()
        return

    print(f"\n  {'Empleado':<16} {'Semanas':<9} {'Mañana':<9} {'Tarde':<9} {'Noche':<9} {'Partido':<9} {'Libre'}")
    print("  " + "-" * 70)
    for nombre in sorted(conteos.keys()):
        c  = conteos[nombre]
        sw = semanas_trabajadas[nombre]
        print(
            f"  {nombre:<16} {sw:<9} "
            f"{c['Mañana']:<9} {c['Tarde']:<9} "
            f"{c['Noche']:<9} {c['Partido']:<9} {c['Libre']}"
        )

    exportar = input("\n  ¿Exportar estadísticas a Excel? (s/n): ").strip().lower()
    if exportar == "s":
        _exportar_estadisticas_excel(conteos, semanas_trabajadas)
    else:
        _pausar()

def _exportar_estadisticas_excel(conteos: dict, semanas_trabajadas: dict):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Estadísticas"

    HEADER_COLOR = "1565C0"
    COL_COLORS   = ["FFF59D", "A5D6A7", "90CAF9", "FFCC80", "EF9A9A"]
    s_thin = Side(style="thin", color="BBBBBB")
    borde  = Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_thin)

    headers = ["Empleado", "Semanas", "Mañana", "Tarde", "Noche", "Partido", "Libre"]
    ws.column_dimensions["A"].width = 18
    for ci in range(2, 8):
        ws.column_dimensions[chr(64 + ci)].width = 11

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = borde
    ws.row_dimensions[1].height = 22

    turnos = ["Mañana", "Tarde", "Noche", "Partido", "Libre"]
    for ri, nombre in enumerate(sorted(conteos.keys()), start=2):
        c  = conteos[nombre]
        sw = semanas_trabajadas[nombre]
        fila = [nombre, sw] + [c[t] for t in turnos]
        for ci, val in enumerate(fila, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")
            cell.border    = borde
            if ci >= 3:
                cell.fill = PatternFill("solid", start_color=COL_COLORS[ci - 3])
        ws.row_dimensions[ri].height = 20

    ruta = os.path.join(BASE_DIR, "estadisticas_turnos.xlsx")
    wb.save(ruta)
    print(f"\n  ✅ Estadísticas exportadas: {ruta}")
    _pausar()

def _eliminar_semana_historial(historial: dict, semanas: list[str]):
    _separador("ELIMINAR SEMANA")
    key = _seleccionar_semana_historial(semanas)
    d   = _semana_desde_str(key)
    confirmar = input(f"  ¿Eliminar semana del {d.strftime('%d/%m/%Y')}? (s/n): ").strip().lower()
    if confirmar == "s":
        del historial[key]
        guardar_historial(historial)
        print("  ✅ Semana eliminada.")
    else:
        print("  Cancelado.")
    _pausar()


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 4 — VACACIONES
# ═══════════════════════════════════════════════════════════════════════════════
#
# Estructura de vacaciones.json:
# [
#   {
#     "id":       "uuid corto",
#     "nombre":   "Pedro",
#     "inicio":   "2026-06-01",   ← primer día de vacaciones (cualquier día)
#     "fin":      "2026-06-14",   ← último día incluido
#     "nota":     "Verano"        ← opcional
#   }, ...
# ]
#
# Al generar el horario de una semana, las vacaciones activas esa semana
# se convierten automáticamente en forced_days_off para los días que solapan.
# ═══════════════════════════════════════════════════════════════════════════════

def cargar_vacaciones(dept_id: str = "recepcion") -> list[dict]:
    return _cargar_json(_dept_file("vacaciones.json", dept_id), [])

def guardar_vacaciones(vacaciones: list[dict], dept_id: str = "recepcion"):
    _guardar_json(_dept_file("vacaciones.json", dept_id), vacaciones)

def _generar_id() -> str:
    import random, string
    return "".join(random.choices(string.ascii_lowercase + string.digits, k=6))

def vacaciones_activas_en_semana(week_start: date, dept_id: str = "recepcion") -> dict:
    dias_semana = [week_start + timedelta(days=i) for i in range(7)]
    vacaciones  = cargar_vacaciones(dept_id)
    resultado: dict[str, list[str]] = {}

    for v in vacaciones:
        try:
            inicio = date.fromisoformat(v["inicio"])
            fin    = date.fromisoformat(v["fin"])
        except (KeyError, ValueError):
            continue
        nombre = v["nombre"]
        for i, dia_date in enumerate(dias_semana):
            if inicio <= dia_date <= fin:
                resultado.setdefault(nombre, [])
                resultado[nombre].append(DIAS_SEMANA[i])

    return resultado

def aplicar_vacaciones(staff_list: list[dict], week_start: date, dept_id: str = "recepcion") -> list[dict]:
    import copy
    vac    = vacaciones_activas_en_semana(week_start, dept_id)
    staff2 = copy.deepcopy(staff_list)
    for person in staff2:
        dias_vac = vac.get(person["name"], [])
        if dias_vac:
            existentes = person.get("forced_days_off", [])
            # Merge sin duplicados, manteniendo orden de la semana
            todos = list(dict.fromkeys(existentes + dias_vac))
            person["forced_days_off"] = todos
    return staff2

def menu_vacaciones():
    """Submenú completo de gestión de vacaciones."""
    while True:
        _limpiar_pantalla()
        _separador("GESTIÓN DE VACACIONES")
        vacaciones = cargar_vacaciones()
        activas    = [v for v in vacaciones if date.fromisoformat(v["fin"]) >= date.today()]
        print(f"  Periodos registrados: {len(vacaciones)}  |  Activos/futuros: {len(activas)}")
        print()
        print("  1. Ver todos los periodos")
        print("  2. Añadir periodo de vacaciones")
        print("  3. Eliminar periodo")
        print("  4. Ver quién está de vacaciones una semana")
        print("  0. Volver")
        opcion = _input_opcion(["1", "2", "3", "4", "0"])

        if opcion == "1":
            _ver_vacaciones(vacaciones)
        elif opcion == "2":
            _añadir_vacaciones(vacaciones)
        elif opcion == "3":
            _eliminar_vacaciones(vacaciones)
        elif opcion == "4":
            _consultar_semana_vacaciones()
        elif opcion == "0":
            break

def _ver_vacaciones(vacaciones: list[dict]):
    _separador("PERIODOS DE VACACIONES")
    if not vacaciones:
        print("  No hay periodos registrados.")
        _pausar()
        return

    hoy = date.today()
    # Ordenar por fecha de inicio
    ordenadas = sorted(vacaciones, key=lambda v: v["inicio"])
    print(f"\n  {'#':<4} {'Empleado':<16} {'Inicio':<13} {'Fin':<13} {'Días':<7} {'Estado':<12} Nota")
    print("  " + "-" * 72)
    for i, v in enumerate(ordenadas, 1):
        ini  = date.fromisoformat(v["inicio"])
        fin  = date.fromisoformat(v["fin"])
        dias = (fin - ini).days + 1
        if fin < hoy:
            estado = "Pasado"
        elif ini <= hoy <= fin:
            estado = "✅ Activo"
        else:
            estado = "Próximo"
        nota = v.get("nota", "—")
        print(f"  {i:<4} {v['nombre']:<16} {ini.strftime('%d/%m/%Y'):<13} "
              f"{fin.strftime('%d/%m/%Y'):<13} {dias:<7} {estado:<12} {nota}")
    _pausar()

def _pedir_fecha(prompt: str, referencia: date = None) -> date:
    """Pide una fecha en formato DD/MM/YYYY o YYYY-MM-DD."""
    while True:
        raw = input(prompt).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                d = date(*[int(x) for x in
                    (raw.split("/") if "/" in raw else raw.split("-"))[::-1 if "/" in raw else 1]])
                # Usar strptime para mayor seguridad
                from datetime import datetime
                d = datetime.strptime(raw, fmt).date()
                if referencia and d < referencia:
                    print(f"  ⚠ La fecha no puede ser anterior a {referencia.strftime('%d/%m/%Y')}.")
                    break
                return d
            except ValueError:
                continue
        print("  ⚠ Formato no reconocido. Usa DD/MM/YYYY o YYYY-MM-DD.")

def _añadir_vacaciones(vacaciones: list[dict]):
    _separador("AÑADIR VACACIONES")
    staff = cargar_plantilla()
    if not staff:
        print("  No hay empleados en la plantilla.")
        _pausar()
        return

    print("\n  Empleados:")
    for i, p in enumerate(staff, 1):
        print(f"    {i}. {p['name']}")
    idx_str = _input_opcion([str(i) for i in range(1, len(staff) + 1)], "  Número de empleado: ")
    nombre  = staff[int(idx_str) - 1]["name"]

    print()
    inicio = _pedir_fecha("  Fecha de inicio (DD/MM/YYYY): ")
    fin    = _pedir_fecha("  Fecha de fin    (DD/MM/YYYY): ", referencia=inicio)
    nota   = input("  Nota (opcional, Enter para omitir): ").strip() or ""

    dias_total = (fin - inicio).days + 1

    # Advertir si solapa con otro periodo del mismo empleado
    solapados = [
        v for v in vacaciones
        if v["nombre"] == nombre
        and date.fromisoformat(v["inicio"]) <= fin
        and date.fromisoformat(v["fin"])    >= inicio
    ]
    if solapados:
        print(f"\n  ⚠ Este periodo solapa con {len(solapados)} periodo(s) existente(s) de {nombre}.")
        continuar = input("  ¿Continuar igualmente? (s/n): ").strip().lower()
        if continuar != "s":
            print("  Cancelado.")
            _pausar()
            return

    nuevo = {
        "id":     _generar_id(),
        "nombre": nombre,
        "inicio": inicio.isoformat(),
        "fin":    fin.isoformat(),
        "nota":   nota,
    }
    vacaciones.append(nuevo)
    guardar_vacaciones(vacaciones)
    semanas_afectadas = (dias_total + 6) // 7
    print(f"\n  ✅ Vacaciones registradas: {nombre}")
    print(f"     {inicio.strftime('%d/%m/%Y')} → {fin.strftime('%d/%m/%Y')}  "
          f"({dias_total} días, ~{semanas_afectadas} semana(s))")
    _pausar()

def _eliminar_vacaciones(vacaciones: list[dict]):
    _separador("ELIMINAR PERIODO DE VACACIONES")
    if not vacaciones:
        print("  No hay periodos registrados.")
        _pausar()
        return
    _ver_vacaciones(vacaciones)
    ordenadas = sorted(vacaciones, key=lambda v: v["inicio"])
    idx_str   = _input_opcion([str(i) for i in range(1, len(ordenadas) + 1)], "  Número a eliminar: ")
    v         = ordenadas[int(idx_str) - 1]
    confirmar = input(f"  ¿Eliminar vacaciones de '{v['nombre']}' "
                      f"({v['inicio']} → {v['fin']})? (s/n): ").strip().lower()
    if confirmar == "s":
        vacaciones.remove(v)
        guardar_vacaciones(vacaciones)
        print("  ✅ Periodo eliminado.")
    else:
        print("  Cancelado.")
    _pausar()

def _consultar_semana_vacaciones():
    _separador("CONSULTAR SEMANA")
    semana = _pedir_semana()
    fin    = semana + timedelta(days=6)
    activas = vacaciones_activas_en_semana(semana)
    print(f"\n  Semana {semana.strftime('%d/%m/%Y')} – {fin.strftime('%d/%m/%Y')}:")
    if not activas:
        print("  Nadie está de vacaciones esta semana.")
    else:
        for nombre, dias in activas.items():
            print(f"    {nombre}: {', '.join(dias)}")
    _pausar()


# ═══════════════════════════════════════════════════════════════════════════════
# MÓDULO 5 — FESTIVOS Y COMPENSACIONES
# ═══════════════════════════════════════════════════════════════════════════════
#
# Festivos oficiales Valencia ciudad = 12 autonómicos/nacionales + 2 locales
# Fuente: Decreto 100/2025 (DOGV 07/07/2025) + Resolución festivos locales 2025
#
# Estructura festivos.json:
# {
#   "YYYY": [
#     {"fecha": "YYYY-MM-DD", "nombre": "Año Nuevo", "tipo": "nacional"},
#     ...
#   ]
# }
#
# Estructura compensaciones.json:
# {
#   "nombre_empleado": {
#     "año_modificacion": 2026,          ← solo se puede cambiar en enero
#     "modalidad": "cobrar"|"librar_mismo_dia"|"dia_libre_eleccion"
#   }
# }
#
# Estructura festivos_trabajados.json:
# {
#   "YYYY-MM-DD": {
#     "festivo": "Nombre del festivo",
#     "trabajadores": {
#       "nombre": "turno"     ← turno que hizo ese día
#     }
#   }
# }
# ═══════════════════════════════════════════════════════════════════════════════

FESTIVOS_FILE       = os.path.join(DATA_DIR, "festivos.json")
COMPENSACIONES_FILE = os.path.join(DATA_DIR, "compensaciones.json")
FEST_TRABAJADOS_FILE= os.path.join(DATA_DIR, "festivos_trabajados.json")

MODALIDADES = {
    "cobrar":            "Cobrar el festivo (plus económico)",
    "librar_mismo_dia":  "Librar ese día exacto en otra semana",
    "dia_libre_eleccion":"Día libre a elección (dentro del año)",
}

# Festivos oficiales Valencia ciudad embebidos por año.
# Cada año se añade una nueva clave. Si el año no está, se avisa al usuario.
FESTIVOS_VALENCIA = {
    2026: [
        {"fecha": "2026-01-01", "nombre": "Año Nuevo",                    "tipo": "nacional"},
        {"fecha": "2026-01-06", "nombre": "Epifanía del Señor",           "tipo": "nacional"},
        {"fecha": "2026-01-22", "nombre": "San Vicente Mártir",           "tipo": "local"},
        {"fecha": "2026-03-19", "nombre": "San José",                     "tipo": "autonómico"},
        {"fecha": "2026-04-03", "nombre": "Viernes Santo",                "tipo": "nacional"},
        {"fecha": "2026-04-06", "nombre": "Lunes de Pascua",              "tipo": "autonómico"},
        {"fecha": "2026-04-13", "nombre": "San Vicente Ferrer",           "tipo": "local"},
        {"fecha": "2026-05-01", "nombre": "Fiesta del Trabajo",           "tipo": "nacional"},
        {"fecha": "2026-06-24", "nombre": "San Juan",                     "tipo": "autonómico"},
        {"fecha": "2026-08-15", "nombre": "Asunción de la Virgen",        "tipo": "nacional"},
        {"fecha": "2026-10-09", "nombre": "Día de la Comunitat Valenciana","tipo": "autonómico"},
        {"fecha": "2026-10-12", "nombre": "Fiesta Nacional de España",    "tipo": "nacional"},
        {"fecha": "2026-12-08", "nombre": "Inmaculada Concepción",        "tipo": "nacional"},
        {"fecha": "2026-12-25", "nombre": "Natividad del Señor",          "tipo": "nacional"},
    ],
    # Añadir aquí años siguientes cuando se publiquen en el DOGV
}

def cargar_festivos(anio: int) -> list[dict]:
    """
    Devuelve la lista de festivos de Valencia ciudad para el año dado.
    Prioriza festivos.json (personalizable) sobre los embebidos en FESTIVOS_VALENCIA.
    """
    datos = _cargar_json(FESTIVOS_FILE, {})
    if str(anio) in datos:
        return datos[str(anio)]
    if anio in FESTIVOS_VALENCIA:
        return FESTIVOS_VALENCIA[anio]
    return []

def es_festivo(d: date) -> dict | None:
    """Devuelve el dict del festivo si la fecha lo es, o None."""
    festivos = cargar_festivos(d.year)
    for f in festivos:
        if date.fromisoformat(f["fecha"]) == d:
            return f
    return None

def festivos_en_semana(week_start: date) -> list[dict]:
    """Devuelve los festivos que caen dentro de la semana dada."""
    resultado = []
    for i in range(7):
        d = week_start + timedelta(days=i)
        f = es_festivo(d)
        if f:
            resultado.append({**f, "dia_semana": DIAS_SEMANA[i], "date": d})
    return resultado

# ── Compensaciones ───────────────────────────────────────────────────────────

def cargar_compensaciones(dept_id: str = "recepcion") -> dict:
    return _cargar_json(_dept_file("compensaciones.json", dept_id), {})

def guardar_compensaciones(comp: dict, dept_id: str = "recepcion"):
    _guardar_json(_dept_file("compensaciones.json", dept_id), comp)

def get_modalidad_empleado(nombre: str) -> str:
    """Devuelve la modalidad de compensación del empleado (default: cobrar)."""
    comp = cargar_compensaciones()
    return comp.get(nombre, {}).get("modalidad", "cobrar")

def menu_compensaciones():
    """Submenú para ver y modificar la modalidad de compensación de festivos."""
    while True:
        _limpiar_pantalla()
        _separador("COMPENSACIÓN DE FESTIVOS POR EMPLEADO")
        staff = cargar_plantilla()
        comp  = cargar_compensaciones()
        hoy   = date.today()

        print(f"\n  {'Empleado':<16} {'Modalidad actual':<38} {'Modificable'}")
        print("  " + "-" * 70)
        for p in staff:
            nombre    = p["name"]
            datos_emp = comp.get(nombre, {})
            modalidad = datos_emp.get("modalidad", "cobrar")
            anio_mod  = datos_emp.get("año_modificacion", None)
            label     = MODALIDADES.get(modalidad, modalidad)
            # Solo modificable en enero
            modificable = "✅ Sí (enero)" if hoy.month == 1 else f"❌ Solo en enero (último cambio: {anio_mod or '—'})"
            print(f"  {nombre:<16} {label:<38} {modificable}")

        print()
        if hoy.month == 1:
            print("  1. Modificar modalidad de un empleado")
        print("  0. Volver")
        opciones = ["1", "0"] if hoy.month == 1 else ["0"]
        opcion = _input_opcion(opciones)

        if opcion == "1":
            _modificar_compensacion(staff, comp, hoy.year)
        elif opcion == "0":
            break

def _modificar_compensacion(staff: list[dict], comp: dict, anio: int):
    _separador("MODIFICAR COMPENSACIÓN")
    print("\n  Empleados:")
    for i, p in enumerate(staff, 1):
        print(f"    {i}. {p['name']}")
    idx_str = _input_opcion([str(i) for i in range(1, len(staff) + 1)], "  Número de empleado: ")
    nombre  = staff[int(idx_str) - 1]["name"]

    datos_emp = comp.get(nombre, {})
    anio_mod  = datos_emp.get("año_modificacion", None)

    if anio_mod == anio:
        print(f"\n  ⚠ {nombre} ya modificó su modalidad en {anio}. Solo se permite un cambio por año.")
        _pausar()
        return

    print(f"\n  Modalidades disponibles:")
    claves = list(MODALIDADES.keys())
    for i, (k, v) in enumerate(MODALIDADES.items(), 1):
        actual = " ← actual" if k == datos_emp.get("modalidad", "cobrar") else ""
        print(f"    {i}. {v}{actual}")

    idx = int(_input_opcion([str(i) for i in range(1, len(claves) + 1)], "  Opción: ")) - 1
    nueva = claves[idx]

    comp[nombre] = {"modalidad": nueva, "año_modificacion": anio}
    guardar_compensaciones(comp)
    print(f"\n  ✅ {nombre}: modalidad actualizada a '{MODALIDADES[nueva]}'")
    _pausar()

# ── Registro de festivos trabajados ─────────────────────────────────────────

def cargar_festivos_trabajados(dept_id: str = "recepcion") -> dict:
    return _cargar_json(_dept_file("festivos_trabajados.json", dept_id), {})

def guardar_festivos_trabajados(datos: dict, dept_id: str = "recepcion"):
    _guardar_json(_dept_file("festivos_trabajados.json", dept_id), datos)

def registrar_festivos_semana(week_start: date, schedule: dict, staff_list: list[dict]):
    """
    Detecta automáticamente qué festivos caen en la semana y qué trabajadores
    estuvieron en turno activo (no Libre) ese día. Se llama desde main.py
    al guardar el historial de una semana.
    """
    festivos = festivos_en_semana(week_start)
    if not festivos:
        return

    datos = cargar_festivos_trabajados()
    DIAS  = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

    for f in festivos:
        fecha_str  = f["date"].isoformat()
        dia_idx    = DIAS.index(f["dia_semana"])
        trabajando = {}
        for person in staff_list:
            turno = schedule.get(person["name"], ["Libre"]*7)[dia_idx]
            if turno != "Libre":
                trabajando[person["name"]] = turno

        # No sobrescribir si ya existe (la primera vez que se genera es canónica)
        if fecha_str not in datos:
            datos[fecha_str] = {
                "festivo":      f["nombre"],
                "tipo":         f["tipo"],
                "trabajadores": trabajando,
            }
        else:
            # En regeneraciones, actualizar solo los trabajadores (el horario pudo cambiar)
            datos[fecha_str]["trabajadores"] = trabajando

    guardar_festivos_trabajados(datos)

def menu_festivos():
    """Menú principal de festivos: conteo, compensaciones y exportación."""
    while True:
        _limpiar_pantalla()
        _separador("FESTIVOS — VALENCIA CIUDAD")
        print("  1. Ver festivos del año")
        print("  2. Ver conteo de festivos trabajados por empleado")
        print("  3. Gestionar modalidad de compensación")
        print("  4. Exportar informe de festivos a Excel")
        print("  0. Volver")
        opcion = _input_opcion(["1", "2", "3", "4", "0"])

        if opcion == "1":
            _ver_festivos_anio()
        elif opcion == "2":
            _ver_conteo_festivos()
        elif opcion == "3":
            menu_compensaciones()
        elif opcion == "4":
            _exportar_festivos_excel()
        elif opcion == "0":
            break

def _ver_festivos_anio():
    _separador("FESTIVOS DEL AÑO")
    anio = date.today().year
    raw  = input(f"  Año [{anio}]: ").strip()
    if raw:
        try:
            anio = int(raw)
        except ValueError:
            print("  ⚠ Año no válido.")
            _pausar()
            return

    festivos = cargar_festivos(anio)
    if not festivos:
        print(f"  No hay festivos registrados para {anio}.")
        print("  Puedes añadirlos manualmente en data/festivos.json")
        _pausar()
        return

    trabajados = cargar_festivos_trabajados()
    print(f"\n  Festivos oficiales Valencia ciudad {anio}:\n")
    print(f"  {'Fecha':<14} {'Día':<12} {'Tipo':<13} {'Nombre':<30} {'Registrado'}")
    print("  " + "-" * 78)
    for f in festivos:
        d         = date.fromisoformat(f["fecha"])
        dia_sem   = DIAS_SEMANA[d.weekday()]
        registrado = "✅" if f["fecha"] in trabajados else "—"
        print(f"  {d.strftime('%d/%m/%Y'):<14} {dia_sem:<12} {f['tipo']:<13} {f['nombre']:<30} {registrado}")
    print(f"\n  Total: {len(festivos)} festivos")
    _pausar()

def _ver_conteo_festivos():
    _separador("CONTEO DE FESTIVOS TRABAJADOS")
    trabajados = cargar_festivos_trabajados()
    comp       = cargar_compensaciones()
    staff      = cargar_plantilla()

    if not trabajados:
        print("  No hay festivos registrados aún. Genera horarios de semanas con festivos.")
        _pausar()
        return

    # Construir conteo por empleado
    conteo: dict[str, list[dict]] = {p["name"]: [] for p in staff}
    for fecha_str, info in sorted(trabajados.items()):
        for nombre, turno in info["trabajadores"].items():
            if nombre in conteo:
                conteo[nombre].append({
                    "fecha":   fecha_str,
                    "festivo": info["festivo"],
                    "tipo":    info["tipo"],
                    "turno":   turno,
                })

    print(f"\n  {'Empleado':<16} {'Festivos':<10} {'Modalidad'}")
    print("  " + "-" * 60)
    for nombre in sorted(conteo.keys()):
        n         = len(conteo[nombre])
        modalidad = MODALIDADES.get(comp.get(nombre, {}).get("modalidad", "cobrar"), "—")
        print(f"  {nombre:<16} {n:<10} {modalidad}")

    ver_detalle = input("\n  ¿Ver detalle de un empleado? (s/n): ").strip().lower()
    if ver_detalle == "s":
        nombres = sorted(conteo.keys())
        for i, n in enumerate(nombres, 1):
            print(f"    {i}. {n}")
        idx_str = _input_opcion([str(i) for i in range(1, len(nombres) + 1)], "  Número: ")
        nombre  = nombres[int(idx_str) - 1]
        print(f"\n  Festivos trabajados por {nombre}:")
        if not conteo[nombre]:
            print("  Ninguno registrado.")
        for entry in conteo[nombre]:
            d = date.fromisoformat(entry["fecha"])
            print(f"    {d.strftime('%d/%m/%Y')}  {entry['festivo']:<28} Turno: {entry['turno']}")
        _pausar()
    else:
        _pausar()

def _exportar_festivos_excel():
    """Exporta un informe completo: festivos × empleado con turno y modalidad de compensación."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    trabajados = cargar_festivos_trabajados()
    comp       = cargar_compensaciones()
    staff      = cargar_plantilla()

    if not trabajados:
        print("  No hay datos de festivos trabajados aún.")
        _pausar()
        return

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Festivos Trabajados"

    HEADER = "1565C0"
    TIPO_COLOR = {"nacional": "FFE0B2", "autonómico": "E1F5FE", "local": "F3E5F5"}
    TURNO_COLOR = {"Mañana": "FFF59D", "Tarde": "A5D6A7", "Noche": "90CAF9",
                   "Partido": "FFCC80"}
    s_thin = Side(style="thin", color="BBBBBB")
    borde  = Border(left=s_thin, right=s_thin, top=s_thin, bottom=s_thin)

    nombres = [p["name"] for p in staff]
    fechas  = sorted(trabajados.keys())

    # Fila 1: título
    total_cols = 3 + len(nombres)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"] = f"Informe Festivos Trabajados — Valencia ciudad  ({date.today().year})"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", start_color=HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Fila 2: cabeceras
    cabeceras = ["Fecha", "Festivo", "Tipo"] + nombres
    for ci, cab in enumerate(cabeceras, 1):
        cell = ws.cell(row=2, column=ci, value=cab)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", start_color=HEADER)
        cell.alignment = Alignment(horizontal="center")
        cell.border    = borde
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 13
    for ci in range(4, 4 + len(nombres)):
        ws.column_dimensions[chr(64 + ci)].width = 13
    ws.row_dimensions[2].height = 20

    # Filas de datos
    for ri, fecha_str in enumerate(fechas, start=3):
        info  = trabajados[fecha_str]
        d     = date.fromisoformat(fecha_str)
        tipo  = info.get("tipo", "nacional")
        color_fila = TIPO_COLOR.get(tipo, "FFFFFF")

        ws.cell(row=ri, column=1, value=d.strftime("%d/%m/%Y")).border = borde
        ws.cell(row=ri, column=1).fill = PatternFill("solid", start_color=color_fila)
        ws.cell(row=ri, column=1).alignment = Alignment(horizontal="center")

        ws.cell(row=ri, column=2, value=info["festivo"]).border = borde
        ws.cell(row=ri, column=2).fill = PatternFill("solid", start_color=color_fila)

        ws.cell(row=ri, column=3, value=tipo).border = borde
        ws.cell(row=ri, column=3).fill = PatternFill("solid", start_color=color_fila)
        ws.cell(row=ri, column=3).alignment = Alignment(horizontal="center")

        for ci, nombre in enumerate(nombres, start=4):
            turno = info["trabajadores"].get(nombre, "Libre")
            cell  = ws.cell(row=ri, column=ci, value=turno)
            cell.border    = borde
            cell.alignment = Alignment(horizontal="center")
            cell.fill      = PatternFill("solid", start_color=TURNO_COLOR.get(turno, "EF9A9A"))
        ws.row_dimensions[ri].height = 18

    # Hoja resumen por empleado
    ws2 = wb.create_sheet("Resumen por Empleado")
    ws2.merge_cells(f"A1:{chr(64 + 4)}1")
    ws2["A1"] = "Resumen festivos trabajados y compensación"
    ws2["A1"].font      = Font(bold=True, size=12, color="FFFFFF")
    ws2["A1"].fill      = PatternFill("solid", start_color=HEADER)
    ws2["A1"].alignment = Alignment(horizontal="center")
    ws2.row_dimensions[1].height = 24

    cab2 = ["Empleado", "Festivos trabajados", "Modalidad compensación", "Pendientes de compensar"]
    for ci, c in enumerate(cab2, 1):
        cell = ws2.cell(row=2, column=ci, value=c)
        cell.font   = Font(bold=True, color="FFFFFF")
        cell.fill   = PatternFill("solid", start_color=HEADER)
        cell.border = borde
        cell.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 22
    ws2.column_dimensions["C"].width = 32
    ws2.column_dimensions["D"].width = 25

    for ri, nombre in enumerate(sorted(nombres), start=3):
        n_trabajados = sum(1 for f in trabajados.values() if nombre in f["trabajadores"])
        modalidad    = MODALIDADES.get(comp.get(nombre, {}).get("modalidad", "cobrar"), "—")
        pendientes   = n_trabajados if comp.get(nombre, {}).get("modalidad", "cobrar") != "cobrar" else 0
        fila = [nombre, n_trabajados, modalidad, pendientes]
        for ci, val in enumerate(fila, 1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.border    = borde
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left")
        ws2.row_dimensions[ri].height = 18

    ws.freeze_panes = "A3"
    ws2.freeze_panes = "A3"

    ruta = os.path.join(BASE_DIR, "informe_festivos.xlsx")
    wb.save(ruta)
    print(f"\n  ✅ Informe exportado: {ruta}")
    _pausar()